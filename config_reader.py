"""
config_reader.py
Reads region/environment configuration from an Excel file.
"""

import openpyxl
from pathlib import Path


class ConfigReader:
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self._data: list[dict] = []
        self._load()

    def _load(self):
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Config file not found: {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1] if cell.value]
        self._data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            record = dict(zip(headers, row))
            if record.get("Country") and record.get("Environment"):
                self._data.append(record)

        wb.close()

    def get_countries(self) -> list[str]:
        return sorted(set(r["Country"] for r in self._data))

    def get_environments(self, country: str) -> list[str]:
        return sorted(set(r["Environment"] for r in self._data if r["Country"] == country))

    def get_config(self, country: str, environment: str) -> dict | None:
        for record in self._data:
            if record["Country"] == country and record["Environment"] == environment:
                return record
        return None

    def reload(self):
        self._load()
